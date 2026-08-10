import pandas as pd
from datetime import datetime
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
import numpy as np

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = BASE_DIR / "templates" / "TIME_&_LABOR_FORM_WE.xlsx"


def FSL(uploaded_file):
    
    df_raw = pd.read_excel(uploaded_file, header=None)
    header_idx = df_raw[df_raw.isin(["Person"]).any(axis=1)].index[0]
    df = pd.read_excel(uploaded_file, header=header_idx)

    df["Person"] = df["Person"].ffill()
    df["Name"] = df["Name"].ffill()

    allowed_paycodes = ["Regular", "OT", "OT Premium 15", "Shift Premium 15"]
    df = df[df["PayCode Name"].isin(allowed_paycodes) & df["Apply Date"].notna()].copy()
    df["Hours"] = (df["Reg Hours"] + df["OT Hours"]).round(2)

    hour_cols = ["Reg Hours", "OT Hours", "3rd Shift"]
    df["ActualHours"] = df[hour_cols].max(axis=1)

    df["T&L"] = ""

    premium = df[df["PayCode Name"].str.contains("Premium", case=False, na=False)]
    premium_employees = premium["Person"].unique()

    premium_keys = set(zip(premium["Person"], premium["Apply Date"], premium["ActualHours"]))

    mask = ((df["PayCode Name"] == "Regular") & (df["Person"].isin(premium_employees)) &
        (~df.apply(lambda x: (x["Person"], x["Apply Date"], x["ActualHours"]) in premium_keys,axis=1)))

    df.loc[mask, "T&L"] = "Eligible"

    df = df[df["Hours"] != 0].copy()

    df["ID"] = df["Person"].astype(str).str.strip()
    df["Employee Name"] = df["Name"].astype(str).str.strip()
    df["Date"] = pd.to_datetime(df["Apply Date"], format="mixed").dt.strftime("%m/%d/%Y")

    df_aggregated = df.groupby(["ID", "Date"], as_index=False).agg(
    {"Reg Hours": "sum", "OT Hours": "sum", "Hours": "sum", "Employee Name": "first", "T&L": "first"})

    max_date = pd.to_datetime(df_aggregated["Date"], format="%m/%d/%Y").max()
    df_aggregated["Weekend"] = pd.offsets.Week(weekday=6).rollforward(max_date).strftime("%m%d%Y")

    df_aggregated = df_aggregated[["ID", "Employee Name", "Date", "Hours", "Weekend", "Reg Hours", "OT Hours", "T&L"]]

    return df_aggregated

#Time and Labor Form

def TimeAndLaborForm(df_customer, query_file):

    df_customer = df_customer[df_customer["T&L"].astype(str).str.contains("eligible", case=False, na=False)].copy()
    df_query = pd.read_excel(query_file, header=1, dtype=str)
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Time & Labor Form"]

    ws["B5"] = "Lisa Viator"
    ws["B6"] = "75897"
    ws["B7"] = "lisaviator@spherion.com"
    ws["B8"] = "385-231-3993 / 801-261-8880"
    ws["E5"] = "FLEX - Salt Lake City, UT"
    ws["E6"] = "4003054001"
    ws["E7"] = datetime.now().strftime("%m/%d/%Y")

    query_col = df_query.columns[22]
    df_customer["ID"] = (df_customer["ID"].astype(str).str.split(".").str[0].str.strip())
    df_query[query_col] = (df_query[query_col].astype(str).str.split(".").str[0].str.strip())

    df_clean = pd.merge(
        df_customer,
        df_query.drop_duplicates(subset=[query_col], keep="first"),
        left_on="ID",
        right_on=query_col,
        how="inner")

    if not df_clean.empty:
        df_tei = pd.DataFrame({
            "A": pd.to_datetime(df_clean["Date"]).dt.strftime("%m/%d/%Y"),
            "B": df_clean["Employee Name_x"],
            "C": df_clean["Employee ID"].astype(str).str.split(".").str[0].str.zfill(11),
            "D": df_clean["Order ID"].astype(str).str.split(".").str[0].str.zfill(11),
            "E": df_clean["Reg Hours"].round(2),
            "F": df_clean["OT Hours"].round(2),
            "G": np.where(df_clean["OT Hours"].round(2) == 0, 18, df_clean["Current Pay Rate"].astype(float),            ),
            "H": np.where(df_clean["OT Hours"].round(2) == 0, 24.48, df_clean["Current Bill Rate"].astype(float)),
            "I": ((df_clean["Reg Hours"].round(2) +  df_clean["OT Hours"].round(2)) * (df_clean["Current Pay Rate"].astype(float))).round(2),
            "K": "BILL",
            "L": "048"})

        for r_idx, row in enumerate(df_tei.itertuples(index=False), start=29):
            for col_name, val in zip(df_tei.columns, row):
                ws[f"{col_name}{r_idx}"] = val

        start_row = 29 + len(df_tei)
    else:
        start_row = 29

    if start_row <= 1000:
        ws.delete_rows(start_row, amount=1000 - start_row + 1)

    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer

# CreateUpload

def generate_tei_output(df_customer, query_file):

    df_customer = df_customer[df_customer["T&L"].isna() | (df_customer["T&L"].astype(str).str.strip() == "")].copy()
    df_query = pd.read_excel(query_file, header=1, dtype=str)

    df_customer['ID'] = df_customer['ID'].astype(str).str.split('.').str[0].str.strip()

    query_id_col = df_query.columns[22]
    df_query[query_id_col] = df_query[query_id_col].astype(str).str.split('.').str[0].str.strip()

    df_query_unique = df_query.drop_duplicates(subset=[query_id_col], keep='first')

    df_merged = pd.merge(
        df_customer,
        df_query_unique, 
        left_on='ID', 
        right_on=query_id_col, 
        how='left',
        suffixes=('', '_query'),
        indicator=True)

    error_mask = df_merged['_merge'] == 'left_only'
    df_errors = df_merged.loc[error_mask, ['ID', 'Employee Name', 'Date', 'Hours']].copy()

    df_clean = df_merged[df_merged['_merge'] == 'both'].copy()

    def process_to_tei(df_input):
        if df_input.empty:
            return pd.DataFrame()
        tei_data = {
            "BRANCH_NUMBER": df_input["Business Unit"].astype(int),
            "PROCESS_DATE": datetime.now().strftime("%m%d%y"),
            "PROCESS_TIME": datetime.now().strftime("%H%M%S"),
            "CUSTOMER_CODE": df_input["Customer ID"].astype(int),
            "TYPE4": "Type4",
            "EMPLOYEE_IDENTIFIER": df_input["Employee ID"].astype(str).str.split('.').str[0].str.zfill(11),
            "ORDER_IDENTIFIER": df_input["Order ID"].astype(int),
            "DAY_DATE": pd.to_datetime(df_input['Date']).dt.strftime('%m%d%Y'),
            "HOURS_DAY": (df_input['Hours'] * 100).round().astype(int),
            "EMPLOYEE_NAME": df_input['Employee Name'],
            "PAY_RATE": 0, "BILL_RATE": 0, "PAY_RATE_OT": 0,
            "BILL_RATE_OT": 0, "PAY_RATE_DT": 0, "BILL_RATE_DT": 0,
            "TIME_RPTG_CD": "", "EARNING_DEDUCTION": "",
            "FLAT_AMOUNT": 0, "MARKUP_PCT": 0, "OTH_EARN_DESCR": "",
            "WEEK_END_DATE": df_input['Weekend']
        }
        return pd.DataFrame(tei_data)

    df_tei = process_to_tei(df_clean)

    return df_tei, df_errors